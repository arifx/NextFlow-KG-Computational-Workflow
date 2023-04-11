from rdflib import Graph, Namespace, Literal
from openpyxl import load_workbook
from rdflib.namespace import RDF, RDFS
import datetime

ex = Namespace('http://purl.com/fsmon#')
g = Graph()
fname='/home/appuser/NextFlow-KG-Computational-Workflow/ENOSE.xlsx'
wb = load_workbook(fname)
sheet = wb['Sheet1']
if 1==1:
    row= list(sheet.rows)[0]
    print(str(row))
#for row in sheet.iter_rows(min_row=2, values_only=True):
      subject_uri = ex['Dataset'] #row[0].value
    subjectsample = ex['Sample']
    subject_name =  fname  #row[1].value
    subject = ex['Dataset']
    g.add((subject, RDF.type, Literal("Dataset" )))
    g.add((subjectsample, ex[f'hasDataset'], subject ))

    subject_name =  fname  #row[1].value
    subject = ex['Dataset']
    g.add((subject, RDF.type, Literal("Dataset" )))
    g.add((ex["Datetime"], RDF.type, Literal(str(datetime.datetime.now() ))))
    g.add((subject, ex['hasDatasetName'], Literal(str(subject_name))))
    g.add(((ex["Location"] ), RDF.type, Literal(str("50.850741597 5.6877722489"))))
    g.add(((ex["SampleType"] ), RDF.type, Literal(str("ChickenFillet"))))
    for i in range(1, len(row)):
            #predicate = ex[f'hasMeasurementType{i-1}']
            predicate = ex[f'hasMeasurementType']
            object_uri = row[i].value
            object = ex[object_uri]
            g.add((subject, predicate, object))

f = open("FoodSafetyMonitoringKG.json", "w")
f.write(g.serialize())
#f.write(g.serialize(format='turtle'))

f.close()
