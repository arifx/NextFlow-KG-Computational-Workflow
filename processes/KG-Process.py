from rdflib import Graph, Namespace, Literal
from openpyxl import load_workbook
from rdflib.namespace import RDF, RDFS
import sys

input_file = sys.argv[1]
output_file = sys.argv[2]

ex = Namespace('http://purl.com/ditect#')
g = Graph()
print("Value of input is: "+input_file)
wb = load_workbook(filename=str(input_file))
sheet = wb['Requested Info']

for row in sheet.iter_rows(min_row=2, values_only=True):
    subject_uri = row[0]
    subject_name = row[1]
    subject = ex[subject_uri]
    g.add((subject, RDF.type, ex[row[0]]))
    g.add((subject, ex['hasName'], Literal(subject_name)))

    for i in range(2, len(row)):
        if row[i]:
            predicate = ex[f'hasConnection{i-1}']
            object_uri = row[i]
            object = ex[object_uri]
            g.add((subject, predicate, object))


g.serialize(destination=output_file, format='json-ld')
