from rdflib import Graph, Namespace, Literal
from openpyxl import load_workbook
from rdflib.namespace import RDF, RDFS

def updateKG(inputKGFilename,inputfile,outputKGfilename):
   ex = Namespace('http://purl.com/ditect#')
   g = Graph()
   wb = load_workbook('test.xlsx')
   sheet = wb['Requested Info']

   for row in sheet.iter_rows(min_row=2, values_only=True):
      subject_uri = row[0]
      subject_name = row[1]
      subject = ex[subject_uri]
      g.add((subject, RDF.type, ex[row[0]]))
      g.add((subject, ex['hasName'], Literal(subject_name)))

      for i in range(2, len(row)):
         if row[i]:
               predicate = ex[f'hasConnection{i-1}']
               object_uri = row[i]
               object = ex[object_uri]
               g.add((subject, predicate, object))

   f = open("FoodSafetyMonitoringKG.json", "w")
   f.write("KG here")
   f.close()


import sys, getopt

def main(argv):
   inputfile = ''
   inputkg
   outputfile = ''
   opts, args = getopt.getopt(argv,"hi:o:",["ifile=","ofile="])
   for opt, arg in opts:
      if opt == '-h':
         print ('test.py -i <inputfile> -o <outputfile>')
         sys.exit()
      elif opt in ("-i", "--ifile"):
         inputfile = arg
      elif opt in ("-o", "--ofile"):
         outputfile = arg
   print ('Input file is ', inputfile)
   print ('Output file is ', outputfile)

if __name__ == "__main__":
   main(sys.argv[1:])