from rdflib import Graph, Namespace, Literal
from openpyxl import load_workbook
from rdflib.namespace import RDF, RDFS

ex = Namespace('http://purl.com/ditect#')
g = Graph()
wb = load_workbook('test.xlsx')
sheet = wb['Requested Info']

for row in sheet.iter_rows(min_row=2, values_only=True):
    subject_uri = row[0]
    subject_name = row[1]
    subject = ex[subject_uri]
    g.add((subject, RDF.type, ex[row[0]]))
    g.add((subject, ex['hasName'], Literal(subject_name)))

    for i in range(2, len(row)):
        if row[i]:
            predicate = ex[f'hasConection{i-1}']
            object_uri = row[i]
            object = ex[object_uri]
            g.add((subject, predicate, object))


f = open("FoodSafetyMonitoringKG.json", "w")
v = g.serialize(format="json-ld").decode('utf-8')
f.write(v)
f.close()

